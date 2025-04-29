from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import random

# Load Excel and add progress tracking
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.active
total_rows = sheet.max_row - 1  # Subtract 1 for header row

# Debug Excel data
print(f"Total data sets to process: {total_rows}")
print(f"Number of columns: {sheet.max_column}")
print("Header row:", [cell.value for cell in sheet[1]])

# ChromeDriver with options
chrome_options = Options()
chrome_options.add_argument('--start-maximized')
driver = webdriver.Chrome(options=chrome_options)
wait = WebDriverWait(driver, 10)

try:
    processed_count = 0
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:  # Skip empty rows
            print(f"Skipping empty row {row_idx}")
            continue
            
        name = row[0]
        # Get answers from Excel row
        radio_answers = [str(row[i]) if row[i] is not None else "" for i in range(1, 6)]
        answer6 = str(row[6]) if len(row) > 6 and row[6] is not None else ""

        # Progress tracking
        processed_count += 1
        print(f"\nProcessing entry {processed_count}/{total_rows} - {name}")
        print(f"Answers: {radio_answers} | Short answer: {answer6}")

        try:
            # Open form with retry mechanism
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    driver.get("https://docs.google.com/forms/d/e/1FAIpQLSediOxEBfo3Zkyf1VlgTfXl9T2xENdAKe9OP-ExEAre9zw2UQ/viewform")
                    time.sleep(2)
                    break
                except Exception as e:
                    if attempt == max_retries - 1:
                        raise
                    print(f"Retry {attempt + 1} loading form...")
                    time.sleep(2)

            # Wait for text inputs and fill name only
            name_input = wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "input.whsOnd.zHQkBf")))
            driver.execute_script("arguments[0].scrollIntoView(true);", name_input)
            time.sleep(0.5)
            name_input.clear()
            name_input.send_keys(name)
            
            # Handle radio button questions (1-5)
            radio_containers = wait.until(EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "div.SG0AAe")))
            
            # Process each radio button question
            for question_idx, container in enumerate(radio_containers[:5]):
                if question_idx < len(radio_answers):
                    try:
                        # Get all options for this question
                        options = container.find_elements(By.CSS_SELECTOR, "div.vd3tt")
                        labels = container.find_elements(By.CSS_SELECTOR, "span.aDTYNe")
                        
                        # Get answer from Excel
                        answer_value = str(radio_answers[question_idx])
                        print(f"Question {question_idx + 1}: Selecting answer {answer_value}")
                        
                        # Find and click matching option
                        for i, label in enumerate(labels):
                            if label.text.strip() == answer_value:
                                driver.execute_script("arguments[0].scrollIntoView(true);", options[i])
                                time.sleep(0.5)
                                options[i].click()
                                time.sleep(0.5)
                                break
                    except Exception as e:
                        print(f"Error on question {question_idx + 1}: {str(e)}")

            # Handle short answer questions (6-7)
            short_answers = wait.until(EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "input.whsOnd.zHQkBf")))
            
            # Fill short answers from Excel
            short_answers[1].send_keys(str(answer6))  # Question 6
            
            # Find and click submit button
            submit_btn = wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "div[role='button'][jsname='M2UYVd']")))
            submit_btn.click()
            
            # Wait for submission confirmation
            wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "div.freebirdFormviewerViewResponseConfirmationMessage")))
            print(f"✓ Successfully submitted form for {name}")
            
            # Add random delay between submissions (2-4 seconds)
            time.sleep(2 + random.random() * 2)
            
        except Exception as e:
            print(f"✗ Error processing {name}: {str(e)}")
            # Save progress to error log
            with open('error_log.txt', 'a', encoding='utf-8') as f:
                f.write(f"Row {row_idx}: {name} - {str(e)}\n")
            continue

    print(f"\nCompleted processing {processed_count} entries")
    print(f"Success rate: {(processed_count/total_rows)*100:.1f}%")

finally:
    driver.quit()
    print("Browser closed")
